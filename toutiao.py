import requests
import json
from openpyxl import Workbook
import time
import hashlib
import os
import datetime
import pandas as pd

start_url = 'https://www.toutiao.com/api/pc/feed/?category=news_hot&utm_source=toutiao&widen=1&max_behot_time='
url = 'https://www.toutiao.com'

headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36'
}
cookies = {'tt_webid': '6925668574955062791'}  # 此处cookies可从浏览器中查找，为了避免被头条禁止爬虫

max_behot_time = '0'   # 链接参数
title = []       # 存储新闻标题
source_url = []  # 存储新闻的链接
s_url = []       # 存储新闻的完整链接
source = []      # 存储发布新闻的公众号
media_url = {}   # 存储公众号的完整链接
comments_count = []  # 存储评论量


def get_as_cp():  # 该函数主要是为了获取as和cp参数，程序参考今日头条中的加密js文件：home_4abea46.js
    zz = {}
    now = round(time.time())
    print(now) # 获取当前计算机时间
    e = hex(int(now)).upper()[2:] #hex()转换一个整数对象为16进制的字符串表示 前两位大写。
    print('e:', e)
    a = hashlib.md5()  #hashlib.md5().hexdigest()创建hash对象并返回16进制结果
    print('a:', a)
    a.update(str(int(now)).encode('utf-8'))
    i = a.hexdigest().upper()
    print('i:', i)
    if len(e)!=8:
        zz = {'as':'479BB4B7254C150',
        'cp':'7E0AC8874BB0985'}
        return zz
    n = i[:5]
    a = i[-5:]
    r = ''
    s = ''
    for i in range(5):
        s= s+n[i]+e[i]
    for j in range(5):
        r = r+e[j+3]+a[j]
    zz ={
    'as':'A1'+s+e[-3:],
    'cp':e[0:3]+r+'E1'
    }
    print('zz:', zz)
    return zz

def getdata(url, headers, cookies):  # 解析网页函数
    r = requests.get(url, headers=headers,cookies=cookies)
    print(url)
    data = json.loads(r.text)
    return data


def savedata(title, s_url, source, media_url,comments_count):  # 存储数据到文件
   
    # 存储数据到xlxs文件
    wb = Workbook()
    if not os.path.isdir(os.getcwd()+'/result'):   # 判断文件夹是否存在
        os.makedirs(os.getcwd()+'/result')  # 新建存储文件夹
    filename = os.getcwd()+'/result/result-' + \
        datetime.datetime.now().strftime('%Y-%m-%d-%H-%m')+'.xlsx'  # 新建存储结果的excel文件
    ws = wb.active
    ws.title = 'data'   # 更改工作表的标题
    ws['A1'] = '标题'   # 对表格加入标题
    ws['B1'] = '新闻链接'
    ws['C1'] = '头条号'
    ws['D1'] = '头条号链接'
    ws['E1'] = '评论量'
    for row in range(2, len(title)+2):   # 将数据写入表格
        _ = ws.cell(column=1, row=row, value=title[row-2])
        _ = ws.cell(column=2, row=row, value=s_url[row-2])
        _ = ws.cell(column=3, row=row, value=source[row-2])
        _ = ws.cell(column=4, row=row, value=media_url[source[row-2]])
        _ = ws.cell(column=5, row=row, value=comments_count[row-2])

    wb.save(filename=filename)  # 保存文件
    
    ###按照评论量进行排序并生成排序后的文件
    data = pd.read_excel(filename)
    data.sort_values(by='评论量', ascending=False, inplace=True)
    data.to_csv('result.csv', encoding='utf-8', index=False,
                header=['title', 's_url', 'source', 'media_url', 'comments_count'])
    print(data)



def main(max_behot_time, title, source_url, s_url, source, media_url,comments_count):   # 主函数
    for i in range(10):   # 此处的数字类似于你刷新新闻的次数，正常情况下刷新一次会出现10条新闻，但也存在少于10条的情况；所以最后的结果并不一定是10的倍数
        ascp = get_as_cp()    # 获取as和cp参数的函数
        
        demo = getdata(start_url+max_behot_time+'&max_behot_time_tmp='+max_behot_time +
                       '&tadrequire=true&as='+ascp['as']+'&cp='+ascp['cp'] + '&_signature=', headers, cookies)
        # print(demo)
        time.sleep(1)
        for j in range(len(demo['data'])):
            # print(demo['data'][j]['title'])
            if demo['data'][j]['title'] not in title:
                title.append(demo['data'][j]['title'])  # 获取新闻标题
                source_url.append(demo['data'][j]['source_url'])  # 获取新闻链接
                source.append(demo['data'][j]['source'])  # 获取发布新闻的公众号

            # if demo['data'][j]['comments_count'] not in comments_count:
            if 'comments_count' in demo['data'][j].keys():
                # demo['data'][j]['comments_count'] = demo['data'][j]['comments_count']
                pass
                
            else:
                demo['data'][j]['comments_count'] = 0

            comments_count.append(int(demo['data'][j]['comments_count'])) ##获取评论量

            if demo['data'][j]['source'] not in media_url:
                if 'media_url' in demo['data'][j].keys():
                    media_url[demo['data'][j]['source']] = url + demo['data'][j]['media_url']  # 获取公众号链接
                else:
                    media_url[demo['data'][j]['source']] = ''
            
        print(max_behot_time)

        # 获取下一个链接的max_behot_time参数的值
        max_behot_time = str(demo['next']['max_behot_time'])

        for index in range(len(title)):
            print('标题：', title[index])
            if 'https' not in source_url[index]:
                s_url.append(url+source_url[index])
                print('新闻链接：', url+source_url[index])
            else:
                print('新闻链接：', source_url[index])
                s_url.append(source_url[index])
                # print('源链接：', url+source_url[index])
            print('头条号：', source[index])
            print(len(title))   # 获取的新闻数量


if __name__ == '__main__':
    main(max_behot_time, title, source_url, s_url, source, media_url,comments_count)
    savedata(title, s_url, source, media_url, comments_count)


# https://www.toutiao.com/hot-event/hot-board/?origin=toutiao_pc&_signature=_02B4Z6wo00901eASNswAAIDAHl5K.1czq7HgNjJAAByE29jPOiCW-Hr4LANhznrLGg6.r9kthaUkWpAC8ZjCtQ3SS6Jf6Ag25oW-alkzU6XKceC1XrxI0ACUSOdeCDAHwnJ1pGIxoA-5olqn9e
